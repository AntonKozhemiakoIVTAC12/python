import openpyxl
import xlsxwriter


def counting_time(sec):
    te = sec // 60
    mi = te % 60
    h = te // 60
    sec = sec % 60
    return f'{int(h):02}:{int(mi):02}:{int(sec):02}'


def seconds(time):
    sec = int(time[-2] + time[-1])
    mn = int(time[-5] + time[-4]) * 60
    hr = int(time[:-6]) * 60 * 60
    return hr + mn + sec


path = "lab3Table data.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

workbook = xlsxwriter.Workbook('lab4Table data.xlsx')
worksheet = workbook.add_worksheet()

for v, prm in [(1, ('B', 'C', 'D', 'E')), (4, ('F', 'G', 'H', 'I')), (8, ('J', 'K', 'L', 'M'))]:

    q_time = [0, 0, 0, 0]
    q_count = [0, 0, 0, 0]
    past_time = 0
    cur_time = 0
    q_time2 = [0, 0, 0, 0]
    params = []
    end_moment = 0
    que_time = [0, 0, 0, 0]
    v_1 = [0, 0, 0, 0]
    v_2 = [0, 0, 0, 0]

    for i in range(2, 102):

        stype = sheet_obj.cell(column=1, row=i).value
        slen = sheet_obj.cell(column=3, row=i).value
        stime = sheet_obj.cell(column=4, row=i).value

        q_count[stype - 1] += 1
        q_time[stype - 1] += slen / v
        v_1[stype - 1] += (slen / v) * 0.01
        v_2[stype - 1] += ((slen / v) ** 2) * 0.01
        cur_time = seconds(stime)

        in_queue = end_moment - cur_time
        if in_queue > 0:
            cur_time += in_queue
            que_time[stype - 1] += in_queue

        params.append((stype, slen, cur_time))

        q_time2[stype - 1] += cur_time - past_time
        past_time = cur_time

        end_moment = slen / v + cur_time

    for i, j in [(1, 'Скорость'), (2, 'Тип'), (3, 'v'), (4, 'v**n'), (5, 'd'), (6, 's'), (7, 'm'), (8, 't'),
                 (9, 'la'),
                 (10, 'p'), (11, 'n'), (12, 'R'), (13, 'L'), (14, 'p'), (15, 'w'), (16, 'u'), (17, 'l'), (18, 'W'),
                 (19, 'U'), (20, 'L')]:
        worksheet.write(f'A{i}', j)

    worksheet.write(f'{prm[0]}1', v)
    R = [0, 0, 0, 0]
    LA = [0, 0, 0, 0]
    la_mass = []
    p_mass = []
    for j, k in ((prm[0], 1), (prm[1], 2), (prm[2], 3), (prm[3], 4)):

        v = q_time[k] / (q_count[k] if q_count[k] != 0 else 1)
        d = (v_2[k] - v_1[k]) ** 2
        s = d ** 0.5
        m = 1 / (v if v != 0 else 1) * 10

        t = q_time2[k] / (q_count[k] if q_count[k] != 0 else 1)
        la = 1 / (t if t != 0 else 1) * 10
        la_mass.append(la)
        p = la / m
        p_mass.append(p)

        n = 1 - p

        for i, q in [(2, k + 1), (3, v), (4, v_2[k]), (5, d), (6, s), (7, m), (8, t), (9, la), (10, p), (11, n)]:
            worksheet.write(f'{j}{i}', q)
    for i in range(4):
        for j in range(i + 1):
            R[i] += p_mass[j]
            LA[i] += la_mass[j]

    worksheet.write(f'{prm[0]}12', R[2])
    worksheet.write(f'{prm[1]}12', R[1])
    worksheet.write(f'{prm[2]}12', R[0])

    worksheet.write(f'{prm[0]}13', LA[2])
    worksheet.write(f'{prm[1]}13', LA[1])
    worksheet.write(f'{prm[2]}13', LA[0])
    W = 0
    U = 0
    L = 0
    W_mass = []
    U_mass = []
    L_mass = []
    for j, k in ((prm[0], 1), (prm[1], 2), (prm[2], 3), (prm[3], 4)):
        t = q_time2[k] / q_count[k]
        la = 1 / t
        p = la / LA[k]
        w = q_time[k] / q_count[k]
        u = (q_count[k] + q_time[k]) / q_count[k]
        l = w * la
        W += p * w
        W_mass.append(W)
        U += p * u
        U_mass.append(U)
        L += la * u
        L_mass.append(L)

        for i, q in [(14, p), (15, w), (16, u), (17, l)]:
            worksheet.write(f'{j}{i}', q)

    rx_idx = 0
    for i in range(2, -1, -1):
        worksheet.write(f'{prm[rx_idx]}18', W_mass[i])
        worksheet.write(f'{prm[rx_idx]}19', U_mass[i])
        worksheet.write(f'{prm[rx_idx]}20', L_mass[i])
        rx_idx += 1

for v, prm in [(1, ('O', 'P', 'Q', 'R')), (4, ('S', 'T', 'U', 'V')), (8, ('W', 'X', 'Y', 'Z'))]:

    queue = [[0, 0, 0, 0]]
    in_queue = 0
    begin = 0
    end = 0

    q_count = [0, 0, 0, 0]
    q_time = [0, 0, 0, 0]
    q_time_2 = [0, 0, 0, 0]
    que_time = [0, 0, 0, 0]
    v_1 = [0, 0, 0, 0]
    v_2 = [0, 0, 0, 0]
    past_time = 0

    for i in range(2, 102):

        stype = sheet_obj.cell(column=1, row=i).value
        slen = sheet_obj.cell(column=3, row=i).value
        stime = sheet_obj.cell(column=4, row=i).value
        cur_time = seconds(stime)

        q_count[stype - 1] += 1
        q_time[stype - 1] += slen / v
        v_1[stype - 1] += (slen / v) * 0.01
        v_2[stype - 1] += ((slen / v) ** 2) * 0.01

        in_queue = end - cur_time
        if in_queue <= 0:
            begin = cur_time
        else:
            que_time[stype - 1] += in_queue
            begin = end

        queue.append([stype, slen, cur_time, begin])

        if stype < queue[-2][0]:
            if cur_time < queue[-2][3]:
                queue[-2], queue[-1] = queue[-1], queue[-2]
                queue[-2][3] -= queue[-1][1]
                queue[-1][3] += queue[-2][1]
                que_time[stype - 1] = que_time[stype - 1] - queue[-1][1] + queue[-2][1]

        q_time_2[stype - 1] += cur_time - past_time
        past_time = cur_time
        end = begin + slen / v

    for i, j in [(1, 'Скорость'), (2, 'Тип'), (3, 'v'), (4, 'v**n'), (5, 'd'), (6, 's'), (7, 'm'), (8, 't'), (9, 'la'),
                 (10, 'p'), (11, 'n'), (12, 'R'), (13, 'L'), (14, 'p'), (15, 'w'), (16, 'u'), (17, 'l'), (18, 'W'),
                 (19, 'U'), (20, 'L')]:
        worksheet.write(f'A{i}', j)

    worksheet.write(f'{prm[0]}1', v)

    R = [0, 0, 0, 0]
    LA = [0, 0, 0, 0]
    la_mass = []
    p_mass = []
    for j, k in ((prm[0], 1), (prm[1], 2), (prm[2], 3), (prm[3], 4)):

        v = q_time[k] / q_count[k]
        d = (v_2[k] - v_1[k]) ** 2
        s = d ** 0.5
        m = 1 / v * 10

        t = q_time2[k] / q_count[k]

        la = 1 / t
        la_mass.append(la)

        p = la / m
        p_mass.append(p)

        n = 1 - p

        for i, q in [(2, k + 1), (3, v), (4, v_2[k]), (5, d), (6, s), (7, m), (8, t), (9, la), (10, p), (11, n)]:
            worksheet.write(f'{j}{i}', q)
    for i in range(4):
        for j in range(i + 1):
            R[i] += p_mass[j]
            LA[i] += la_mass[j]

    worksheet.write(f'{prm[0]}12', R[2])
    worksheet.write(f'{prm[1]}12', R[1])
    worksheet.write(f'{prm[2]}12', R[0])

    worksheet.write(f'{prm[0]}13', LA[2])
    worksheet.write(f'{prm[1]}13', LA[1])
    worksheet.write(f'{prm[2]}13', LA[0])
    W = 0
    U = 0
    L = 0
    W_mass = []
    U_mass = []
    L_mass = []
    for j, k in ((prm[0], 1), (prm[1], 2), (prm[2], 3), (prm[3], 4)):
        t = q_time2[k] / q_count[k]
        la = 1 / t
        p = la / LA[k]
        w = q_time[k] / q_count[k]
        u = (q_count[k] + q_time[k]) / q_count[k]
        l = w * la
        W += p * w
        W_mass.append(W)
        U += p * u
        U_mass.append(U)
        L += la * u
        L_mass.append(L)

        for i, q in [(14, p), (15, w), (16, u), (17, l)]:
            worksheet.write(f'{j}{i}', q)

    rx_idx = 0
    for i in range(2, -1, -1):
        worksheet.write(f'{prm[rx_idx]}18', W_mass[i])
        worksheet.write(f'{prm[rx_idx]}19', U_mass[i])
        worksheet.write(f'{prm[rx_idx]}20', L_mass[i])
        rx_idx += 1

workbook.close()

workbook = xlsxwriter.Workbook('lab4 data2.xlsx')
worksheet = workbook.add_worksheet()

worksheet.write_row('A1', ['Номер заявки', 'Момент  появления ее в канале', 'Момент начала  обслуживания',
                           'Момент конца обслуживания', 'Время ожидания', 'Время пребывания в канале'])

queue = queue[1:]
for i in range(100):
    end = counting_time(queue[i][3] + queue[i][1])
    wait = queue[i][3] - queue[i][2]
    c = counting_time(wait + queue[i][1])
    worksheet.write_row(f'A{i + 2}',
                        [i, counting_time(queue[i][2]), counting_time(queue[i][3]), end, counting_time(wait), c])

workbook.close()
